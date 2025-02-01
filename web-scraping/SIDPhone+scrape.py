import os
import requests
import pdfkit
from datetime import datetime
from bs4 import BeautifulSoup
import base64
import pandas as pd
from io import StringIO
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter

#path
#WKHTMLTOPDF_PATH = r"D:\Elam\Sementara\sid\wkhtmltopdf\bin\wkhtmltopdf.exe"
WKHTMLTOPDF_PATH = r"D:\sementara\applications\wkhtmltopdf\bin\wkhtmltopdf.exe"
config = pdfkit.configuration(wkhtmltopdf=WKHTMLTOPDF_PATH)

# Konfigurasi API
API_BASE_URL = "https://demo.sekolahan.id/api"
BEARER_TOKEN = "eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJzbWstY2FrcmEtbnVzYW50YXJhLnNla29sYWhhbi5pZCIsImF1ZCI6IjEwLjEzMC40Ni41OCIsImlhdCI6MTcwNDYyMDU1NSwibmJmIjoxNzA0NjIwNTY1LCJkYXRhIjp7ImlkIjpudWxsfX0._9Geu5biEBUJ6jf89FtuINcP1rDcPHZ0t9vOAQN1hZk"
HEADERS = {"Authorization": f"Bearer {BEARER_TOKEN}"}

# Fungsi untuk mendapatkan JSON response dari API
def get_json_response(url, method="GET", data=None):
    """Mengambil data JSON dari API dengan metode GET atau POST."""
    if method == "POST":
        response = requests.post(url, headers=HEADERS, data=data)
    else:
        response = requests.get(url, headers=HEADERS)
    
    if response.status_code == 200:
        return response.json().get("responseData", {}).get("results", [])
    print(f"Gagal mengambil data dari {url}. Status code: {response.status_code}")
    return None

# Fungsi untuk mencari sekolah
def cari_sekolah():
    """Mencari sekolah berdasarkan nama dan mengembalikan ID sekolah dan subdomain."""
    while True:
        nama_sekolah = input("Masukkan nama sekolah: ")
        sekolah_url = f"{API_BASE_URL}/sekolahdata?namasekolah={nama_sekolah}"
        sekolah_list = get_json_response(sekolah_url)

        if sekolah_list:
            for index, sekolah in enumerate(sekolah_list, start=1):
                # TAMPILKAN NAMA SEKOLAH TANPA KOMA (jika ada)
                cleaned_nama = sekolah['nama'].split(',')[0].strip()
                print(f"{index}. {cleaned_nama}")
            try:
                pilihan = int(input("Pilih sekolah (nomor): "))
                id_sekolah = sekolah_list[pilihan - 1]["id"]
                nama_sekolah = sekolah_list[pilihan - 1]["nama"]
                original_nama = sekolah_list[pilihan - 1]["nama"]
                sanitized_nama = original_nama.replace(',', '').strip()  # Hapus koma
                identifier = sekolah_list[pilihan - 1]["identifier"]
                subdomain = base64.b64decode(identifier).decode('utf-8')  # Decode base64 untuk mendapatkan subdomain
                return id_sekolah, sanitized_nama, subdomain
            except (ValueError, IndexError):
                print("Pilihan tidak valid. Coba lagi.")
        else:
            print("Sekolah tidak ditemukan.")

# Fungsi untuk mengambil data kelas
def get_kelas(id_sekolah):
    """Mengambil daftar kelas dari sekolah."""
    kelas_url = f"{API_BASE_URL}/{id_sekolah}/datakelas"
    return get_json_response(kelas_url)

# Fungsi untuk mengambil data siswa dalam kelas
def get_siswa(id_sekolah, kelas_id):
    """Mengambil daftar siswa dalam suatu kelas."""
    siswa_url = f"{API_BASE_URL}/v2/{id_sekolah}/listsiswakelas"
    data_siswa = get_json_response(siswa_url, "POST", {"idkelas": kelas_id, "idsiswa": ""})
    return data_siswa

# Fungsi untuk mengambil profil siswa
def get_profil_siswa(id_sekolah, idsiswa):
    """Mengambil profil siswa berdasarkan ID siswa."""
    profil_url = f"{API_BASE_URL}/v2/{id_sekolah}/profilsiswa"
    return get_json_response(profil_url, "POST", {"idsiswa": idsiswa})

# Fungsi untuk login ke subdomain
def login(subdomain):
    """Login ke subdomain untuk mendapatkan sesi login."""
    login_url = f'https://{subdomain}.sekolahan.id/login/proses'
    main_url = f'https://{subdomain}.sekolahan.id/'
    login_data = {
        'username': 'superadmin',
        'password': 'sigarantang',
        'submit': ''
    }

    session = requests.Session()
    headers = {'User-Agent': 'Mozilla/5.0'}
    session.get(main_url, headers=headers)

    login_response = session.post(login_url, data=login_data, headers=headers)
    if login_response.status_code == 200:
        return session
    print("Login gagal.")
    return None

# Fungsi untuk scrape data siswa
def scrape_siswa(session, subdomain, idsiswa):
    """Scraping data siswa dari halaman edit."""
    data_siswa_url = f'https://{subdomain}.sekolahan.id/datasiswa/edit/{idsiswa}'
    response = session.get(data_siswa_url)
    if response.status_code == 200:
        soup = BeautifulSoup(response.text, 'html.parser')
        fields = ['nama', 'nik', 'no_kk', 'ayah_nik', 'ibu_nik', 'wali_nik', 'no_kip', 'nm_kip', 'no_kps', 'no_kks', 'tglditerima', 'asalsekolah']
        data = {field: get_input_value(soup, field) for field in fields}
        return data
    else:
        print(f"Gagal mengakses halaman siswa, status code: {response.status_code}")
    return {}

# Tambahkan fungsi ini setelah fungsi scrape_siswa
def download_pdf(session, subdomain, idsiswa, output_path):
    """Download HTML via session, lalu konversi ke PDF"""
    pdf_url = f'https://{subdomain}.sekolahan.id/studentrecord/printdata/{idsiswa}'
    response = session.get(pdf_url)
    
    if response.status_code == 200:
        try:
            # Perbaiki URL relatif dalam HTML
            html_content = response.text.replace('src="/', f'src="https://{subdomain}.sekolahan.id/')
            html_content = html_content.replace('href="/', f'href="https://{subdomain}.sekolahan.id/')
            
            # Simpan cookie session untuk wkhtmltopdf
            cookies = session.cookies.get_dict()
            cookies_str = "; ".join([f"{k}={v}" for k, v in cookies.items()])
            
            options = {
                'quiet': '',
                'print-media-type': '',
                'encoding': 'UTF-8',
                'cookie': [('cookie_name', cookies_str)],  # Kirim cookie
                'custom-header': [
                    ('User-Agent', 'Mozilla/5.0'),  # Samakan dengan header browser
                ],
                'enable-local-file-access': ''  # Izinkan akses file lokal
            }
            
            # Konversi HTML string ke PDF
            pdfkit.from_string(
                html_content, 
                output_path, 
                options=options, 
                configuration=config
            )
            return True
        except Exception as e:
            print(f"Gagal konversi PDF: {str(e)}")
            return False
    else:
        print(f"Gagal akses halaman. Status code: {response.status_code}")
    return False

# Fungsi untuk mendapatkan nilai input dari form
def get_input_value(soup, field):
    """Mengambil nilai dari input berdasarkan nama field."""
    input_tag = soup.find('input', {'name': field})
    return input_tag['value'] if input_tag else None

# TAMBAHKAN FUNGSI INI SETELAH FUNGSI get_input_value
def scrape_profil_sekolah(session, subdomain, sanitized_nama_sekolah):
    """Scraping dan menyimpan profil sekolah ke file txt."""
    PROFIL_SEKOLAH_URL = f'https://{subdomain}.sekolahan.id/profilsekolah'  # Ganti URL sesuai kebutuhan
    response = session.get(PROFIL_SEKOLAH_URL)
    
    if response.status_code != 200:
        print(f"Gagal mengakses profil sekolah. Status code: {response.status_code}")
        return

    soup = BeautifulSoup(response.text, 'html.parser')
    # Profil Sekolah
    # nama_sekolah = get_input_value(soup, 'pnamasekolah')
    original_nama_sekolah = get_input_value(soup, 'pnamasekolah')
    nss = get_input_value(soup, 'nsssekolah')
    npsn = get_input_value(soup, 'npsnsekolah')
    alamat = get_input_value(soup, 'alamat_sekolah')
    kode_pos = get_input_value(soup, 'kode_possekolah')
    desa_kelurahan = get_input_value(soup, 'desa_kelurahansekolah')
    kecamatan = get_input_value(soup, 'kecamatan_sekolah')
    kabupaten_kota = get_input_value(soup, 'kabupatenkota_sekolah')
    provinsi = get_input_value(soup, 'provinsi_sekolah')
    latitude = get_input_value(soup, 'latitude')
    longitude = get_input_value(soup, 'longitude')

    # Informasi Sekolah
    nomor_telepon = get_input_value(soup, 'nomor_teleponsekolah')
    nomor_fax = get_input_value(soup, 'nomor_faxsekolah')
    email = get_input_value(soup, 'emailsekolah')
    website = get_input_value(soup, 'websitesekolah')
    # HANDLE STATUS KEPEMILIKAN DENGAN CHECK BERLAPIS
    # status_kepemilikan = soup.find('select', {'name': 'status_pemilik'}).find('option', selected=True).text if soup.find('select', {'name': 'status_pemilik'}) else 'Tidak ditemukan'
    status_kepemilikan = 'Tidak ditemukan'
    select_status = soup.find('select', {'name': 'status_pemilik'})
    if select_status:
        selected_option = select_status.find('option', selected=True)
        status_kepemilikan = selected_option.text.strip() if selected_option else ''

    # Kelengkapan Sekolah
    sk_pendirian = get_input_value(soup, 'sk_pendirian_sekolah')
    tgl_sk_pendirian = get_input_value(soup, 'tgl_sk_pendirian_sekolah')
    sk_izin_operasional = get_input_value(soup, 'sk_izin_operasional')
    tgl_sk_izin_operasional = get_input_value(soup, 'tgl_sk_izin_operasional')
    no_rekening = get_input_value(soup, 'no_rekeningsekolah')
    nama_bank = get_input_value(soup, 'nama_banksekolah')
    rekening_atas_nama = get_input_value(soup, 'rekening_atas_nama')

    # Profile Yayasan
    nama_yayasan = get_input_value(soup, 'namayayasan')
    pimpinan_yayasan = get_input_value(soup, 'nama_pimpinan_yayasan')
    alamat_yayasan = get_input_value(soup, 'alamat_yayasan')
    kode_pos_yayasan = get_input_value(soup, 'kode_posyayasan')
    desa_kelurahan_yayasan = get_input_value(soup, 'desa_kelurahanyayasan')
    sk_pendirian_yayasan = get_input_value(soup, 'sk_pendirian_yayasan')
    tgl_sk_pendirian_yayasan = get_input_value(soup, 'tgl_sk_pendirian_yayasan')

        # ========== SIMPAN KE FILE ==========
    output_folder = os.path.join("Data Sekolah", sanitized_nama_sekolah)
    os.makedirs(output_folder, exist_ok=True)
    
    output_path = os.path.join(output_folder, "informasi sekolah.txt")
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write("=== PROFIL SEKOLAH ===\n")
        f.write(f"Nama Sekolah     : {original_nama_sekolah}\n")
        f.write(f"NSS              : {nss}\n")
        f.write(f"NPSN             : {npsn}\n")
        f.write(f"Alamat           : {alamat}\n")
        f.write(f"Kode Pos         : {kode_pos}\n")
        f.write(f"Desa/Kelurahan   : {desa_kelurahan}\n")
        f.write(f"Kecamatan        : {kecamatan}\n")
        f.write(f"Kabupaten/Kota   : {kabupaten_kota}\n")
        f.write(f"Provinsi         : {provinsi}\n")
        f.write(f"Latitude         : {latitude}\n")
        f.write(f"Longitude        : {longitude}\n\n")
        
        f.write("=== INFORMASI KONTAK ===\n")
        f.write(f"Telepon          : {nomor_telepon}\n")
        f.write(f"Fax              : {nomor_fax}\n")
        f.write(f"Email            : {email}\n")
        f.write(f"Website          : {website}\n")
        f.write(f"Status Kepemilikan: {status_kepemilikan}\n\n")
        
        f.write("=== LEGALITAS ===\n")
        f.write(f"SK Pendirian     : {sk_pendirian}\n")
        f.write(f"Tanggal SK       : {tgl_sk_pendirian}\n")
        f.write(f"SK Izin Operasional: {sk_izin_operasional}\n")
        f.write(f"Tanggal SK Izin  : {tgl_sk_izin_operasional}\n\n")
        
        f.write("=== REKENING SEKOLAH ===\n")
        f.write(f"Nomor Rekening   : {no_rekening}\n")
        f.write(f"Nama Bank        : {nama_bank}\n")
        f.write(f"Atas Nama        : {rekening_atas_nama}\n\n")
        
        f.write("=== YAYASAN ===\n")
        f.write(f"Nama Yayasan     : {nama_yayasan}\n")
        f.write(f"Pimpinan         : {pimpinan_yayasan}\n")
        f.write(f"Alamat Yayasan   : {alamat_yayasan}\n")
        f.write(f"Kode Pos Yayasan : {kode_pos_yayasan}\n")
        f.write(f"Desa/Kelurahan   : {desa_kelurahan_yayasan}\n")
        f.write(f"Akte Pendirian   : {sk_pendirian_yayasan}\n")
        f.write(f"Tanggal Akte     : {tgl_sk_pendirian_yayasan}\n")

    print(f"\nProfil sekolah berhasil disimpan di: {output_path}")

# Download data alumni
def download_alumni(session, subdomain, nama_sekolah):
    alumni_url = f'https://{subdomain}.sekolahan.id/dataalumni/cetakalumni/'
    
    try:
        response = session.get(alumni_url)
        
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            table = soup.find('table')
            
            if not table:
                print("Tabel alumni tidak ditemukan di halaman.")
                return False
            
            # PERBAIKAN 1: Gunakan StringIO untuk handle HTML
            html_content = str(table)
            df = pd.read_html(StringIO(html_content))[0]  # ← Pakai StringIO
            
            # PERBAIKAN 2: Simpan sebagai .xlsx (format modern)
            output_folder = os.path.join("Data Sekolah", nama_sekolah, "Data Siswa")
            os.makedirs(output_folder, exist_ok=True)
            output_path = os.path.join(output_folder, "alumni.xlsx")  # Ganti ke .xlsx

            # Simpan ke Excel dengan styling
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Alumni')
                
                # Ambil worksheet
                workbook = writer.book
                worksheet = writer.sheets['Alumni']
                
                # Tentukan border
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Terapkan border ke semua cell
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Handle kolom NIK dan HP (tambahkan ` di depan)
                for col in df.columns:
                    if 'NIK' in col or 'HP' in col:
                        col_idx = df.columns.get_loc(col) + 1  # Kolom Excel dimulai dari 1
                        col_letter = get_column_letter(col_idx)
                        
                        for cell in worksheet[col_letter]:
                            if cell.value:  # Jika ada isinya
                                cell.value = f"`{cell.value}"  # Tambahkan ` di depan
                
                # Auto adjust lebar kolom
                for col in worksheet.columns:
                    max_length = max(len(str(cell.value)) for cell in col)
                    adjusted_width = (max_length + 2) * 1.2  # Tambahkan sedikit padding
                    worksheet.column_dimensions[col[0].column_letter].width = adjusted_width
            
            print(f"\nData alumni berhasil disimpan: {output_path}")
            return True
            
        else:
            print(f"Gagal download alumni. Status code: {response.status_code}")
            return False
            
    except Exception as e:
        print(f"Error saat download alumni: {str(e)}")
        return False

#download data guru
def download_guru(session, subdomain, nama_sekolah):
    """Mendownload data guru dalam format XLSX"""
    guru_url = f'https://{subdomain}.sekolahan.id/dataguru/cetakguru/'
    
    try:
        response = session.get(guru_url)
        
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            table = soup.find('table')
            
            if not table:
                print("Tabel guru tidak ditemukan di halaman.")
                return False
            
            # Konversi tabel HTML ke DataFrame
            html_content = str(table)
            df = pd.read_html(StringIO(html_content))[0]
            
            # Path penyimpanan
            output_folder = os.path.join("Data Sekolah", nama_sekolah, "Data Guru")
            os.makedirs(output_folder, exist_ok=True)
            output_path = os.path.join(output_folder, "Data Guru.xlsx")
            
            # Simpan ke Excel dengan styling
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Guru')
                
                # Ambil worksheet
                workbook = writer.book
                worksheet = writer.sheets['Guru']
                
                # Tentukan border
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Terapkan border dan alignment
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Handle kolom NIK/HP
                for col in df.columns:
                    if any(k in col for k in ['NIK', 'HP', 'Telp', 'No. Telpon']):
                        col_idx = df.columns.get_loc(col) + 1
                        col_letter = get_column_letter(col_idx)
                        
                        for cell in worksheet[col_letter]:
                            if cell.value:
                                cell.value = f"`{cell.value}"
                
                # Auto adjust lebar kolom
                for col in worksheet.columns:
                    max_length = max(len(str(cell.value)) for cell in col)
                    adjusted_width = (max_length + 2) * 1.2
                    worksheet.column_dimensions[col[0].column_letter].width = adjusted_width
            
            print(f"\nData guru berhasil disimpan: {output_path}")
            return True
            
        else:
            print(f"Gagal download guru. Status code: {response.status_code}")
            return False
            
    except Exception as e:
        print(f"Error saat download guru: {str(e)}")
        return False

def scrape_guru(session, subdomain, nama_sekolah):
    """Scraping data guru dan download profil PDF."""
    DATA_GURU_URL = f'https://{subdomain}.sekolahan.id/dataguru'

    def scrape_page(url):
        """Scrape data guru dari halaman tertentu."""
        response = session.get(url)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            table = soup.find('table', class_='table table-striped table-hover')
            if table:
                rows = table.find('tbody').find_all('tr')
                guru_data = []
                for row in rows:
                    cols = row.find_all('td')
                    if len(cols) > 1:
                        id_guru = cols[2].text.strip()
                        nama_guru = cols[3].find('strong')
                        nama_guru = nama_guru.text.strip() if nama_guru else cols[3].text.strip()
                        guru_data.append((id_guru, nama_guru))
                return guru_data
        print(f"Gagal mengakses halaman: {url}, status code: {response.status_code}")
        return []

    def download_profil_guru(id_guru, nama_guru):
        """Download profil guru sebagai PDF dengan mengonversi HTML ke PDF."""
        # Buat URL lengkap untuk profil guru
        profil_url = f'https://{subdomain}.sekolahan.id/dataguru/cetakprofil/{id_guru}'
        
        # Print URL lengkap yang sedang diakses
        print(f"Mengakses URL untuk profil guru: {profil_url}")
        
        # Menggunakan session yang sudah login untuk ambil HTML
        response = session.get(profil_url)

        if response.status_code == 200:
            # Print HTML yang diterima untuk memverifikasi
            # print(f"HTML diterima untuk {nama_guru}:\n{response.text[:500]}...")  # Cukup tampilkan 500 karakter pertama

            # Path penyimpanan
            output_folder = os.path.join("Data Sekolah", nama_sekolah, "data guru")
            os.makedirs(output_folder, exist_ok=True)
            output_path = os.path.join(output_folder, f"{nama_guru}.pdf")

            try:
                # Mengonversi HTML ke PDF dengan beberapa parameter tambahan
                options = {
                    'no-images': '',  # Matikan gambar agar tidak dimuat
                    'enable-local-file-access': '',  # Akses file lokal (misalnya resource CSS/JS)
                    'javascript-delay': '2000',  # Memberi waktu lebih untuk memuat elemen dinamis
                    'load-error-handling': 'ignore',  # Mengabaikan error yang muncul saat memuat resource
                }

                # Mengonversi HTML ke PDF dan menyimpannya dengan konfigurasi wkhtmltopdf
                pdfkit.from_string(response.text, output_path, configuration=config, options=options)
                print(f"Profil {nama_guru} berhasil disimpan sebagai PDF: {output_path}")
                return True
            except Exception as e:
                print(f"Gagal mengonversi HTML ke PDF untuk {nama_guru}: {e}")
                return False
        else:
            print(f"Gagal download profil {nama_guru}. Status code: {response.status_code}")
            return False

    # Ambil data guru dari semua halaman
    first_page_response = session.get(DATA_GURU_URL)
    if first_page_response.status_code == 200:
        soup = BeautifulSoup(first_page_response.text, 'html.parser')
        pagination = soup.find('ul', class_='pagination')
        total_pages = max([int(a.text) for a in pagination.find_all('a') if a.text.isdigit()], default=1)

        print(f"Total halaman: {total_pages}")

        all_guru_data = []
        for page in range(1, total_pages + 1):
            page_url = f"{DATA_GURU_URL}/?halaman={page}"
            print(f"Mengakses halaman: {page_url}")
            all_guru_data.extend(scrape_page(page_url))

        # Download profil guru
        for id_guru, nama_guru in all_guru_data:
            print(f"\nMendownload profil: {nama_guru} (ID: {id_guru})")
            download_profil_guru(id_guru, nama_guru)
    else:
        print(f"Gagal mengakses halaman guru, status code: {first_page_response.status_code}")

# Fungsi untuk menyimpan data siswa ke dalam file terpisah
# Modifikasi fungsi simpan_data_siswa (tambahkan bagian untuk download PDF)
def simpan_data_siswa(nama_sekolah, kelas, siswa_data, id_sekolah, session, subdomain):
    """Menyimpan data siswa dan profilnya ke dalam file terpisah berdasarkan nama siswa dan kelas."""
    output_folder = os.path.join(os.getcwd(), "Data Sekolah", nama_sekolah, "Data Siswa", kelas)
    os.makedirs(output_folder, exist_ok=True)
    
    for siswa in siswa_data:
        # Mengambil nama siswa dari profil siswa
        profil = get_profil_siswa(id_sekolah, siswa['idsiswa'])
        nama_siswa = profil.get('nama', 'Nama Tidak Ditemukan')
        siswa_folder = os.path.join(output_folder, nama_siswa)
        os.makedirs(siswa_folder, exist_ok=True)
        
        # Simpan data txt
        file_name_txt = f"{nama_siswa}.txt"
        output_file_path_txt = os.path.join(siswa_folder, file_name_txt)
        
        with open(output_file_path_txt, "w") as file:
            file.write(f"=== Data Siswa: {nama_siswa} ===\n")
            file.write(f"ID: {siswa['idsiswa']}\n")
            
            if 'kelas' in siswa:
                file.write(f"Kelas: {siswa['kelas']}\n")
            else:
                file.write(f"Kelas: Data Tidak Tersedia\n")
            
            file.write("\n")
            
            tgllahir = profil.get("tgllahir", "")
            password = datetime.strptime(tgllahir, "%Y-%m-%d").strftime("%y%m%d") if tgllahir and tgllahir != '0000-00-00' else "default_password"
            file.write(f"Password: {password}\n")
            
            scraped_data = scrape_siswa(session, subdomain, siswa['idsiswa'])
            for key, value in {**profil, **scraped_data}.items():
                file.write(f"{key}: {value}\n")
            file.write("\n")
        
        # Download PDF
        file_name_pdf = f"{nama_siswa}.pdf"
        output_file_path_pdf = os.path.join(siswa_folder, file_name_pdf)
        
        if download_pdf(session, subdomain, siswa['idsiswa'], output_file_path_pdf):
            print(f"PDF untuk {nama_siswa} berhasil didownload")
        else:
            print(f"Gagal download PDF untuk {nama_siswa}")
        
        print(f"Data untuk {nama_siswa} berhasil disimpan dalam {file_name_txt}")

# Main Program
def main():
    id_sekolah, nama_sekolah, subdomain = cari_sekolah()
    # kelas_list = get_kelas(id_sekolah)
    
    # if not kelas_list:
    #     print("Gagal mendapatkan data kelas.")
    #     return

    # Login ke subdomain
    session = login(subdomain)
    if session is None:
        return    

    scrape_profil_sekolah(session, subdomain, nama_sekolah)

    download_alumni(session, subdomain, nama_sekolah)

    download_guru(session, subdomain, nama_sekolah)

    scrape_guru(session, subdomain, nama_sekolah)

    # for kelas in kelas_list:
    #     print(f"Mengambil data siswa untuk kelas: {kelas['namakelas']}")
    #     siswa_list = get_siswa(id_sekolah, kelas["kelasid"])
        
    #     if siswa_list:
    #         simpan_data_siswa(nama_sekolah, kelas['namakelas'], siswa_list, id_sekolah, session, subdomain)
    #         print(f"Data siswa kelas {kelas['namakelas']} berhasil disimpan.")
    #     else:
    #         print(f"Tidak ada data siswa di kelas {kelas['namakelas']}")

if __name__ == "__main__":
    main()