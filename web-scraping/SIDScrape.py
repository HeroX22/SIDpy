import os
import requests
import pdfkit
from datetime import datetime
from bs4 import BeautifulSoup
import base64
import pandas as pd
from io import StringIO
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter

#path
#WKHTMLTOPDF_PATH = r"D:\Elam\Sementara\sid\wkhtmltopdf\bin\wkhtmltopdf.exe" # untuk laptop utama
WKHTMLTOPDF_PATH = r"D:\sementara\applications\wkhtmltopdf\bin\wkhtmltopdf.exe" #laptop kedua
#WKHTMLTOPDF_PATH = r"/usr/local/bin/wkhtmltopdf " #linux
config = pdfkit.configuration(wkhtmltopdf=WKHTMLTOPDF_PATH)

# Konfigurasi API
API_BASE_URL = "https://demo.sekolahan.id/api"
BEARER_TOKEN = "eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJzbWstY2FrcmEtbnVzYW50YXJhLnNla29sYWhhbi5pZCIsImF1ZCI6IjEwLjEzMC40Ni41OCIsImlhdCI6MTcwNDYyMDU1NSwibmJmIjoxNzA0NjIwNTY1LCJkYXRhIjp7ImlkIjpudWxsfX0._9Geu5biEBUJ6jf89FtuINcP1rDcPHZ0t9vOAQN1hZk"
HEADERS = {"Authorization": f"Bearer {BEARER_TOKEN}"}

# Fungsi untuk membaca daftar nama sekolah dari file
def baca_daftar_sekolah(file_path):
    """Membaca daftar nama sekolah dari file dan mengembalikan list."""
    with open(file_path, 'r', encoding='utf-8') as file:
        return [line.strip() for line in file if line.strip()]

# Fungsi untuk mendapatkan JSON response dari API
def get_json_response(url, method="GET", data=None):
    """Mengambil data JSON dari API dengan metode GET atau POST."""
    if method == "POST":
        response = requests.post(url, headers=HEADERS, data=data)
    else:
        response = requests.get(url, headers=HEADERS)
    
    if response.status_code == 200:
        return response.json().get("responseData", {}).get("results", [])
    print(f"Gagal mengambil data dari {url}. Status code: {response.status_code}")
    return None

# Fungsi untuk mencari sekolah
def cari_sekolah(daftar_nama_sekolah):
    """Mencari sekolah berdasarkan daftar nama dan mengembalikan ID sekolah dan subdomain."""
    for nama_sekolah in daftar_nama_sekolah:
        sekolah_url = f"{API_BASE_URL}/sekolahdata?namasekolah={nama_sekolah}"
        sekolah_list = get_json_response(sekolah_url)

        if sekolah_list:
            for sekolah in sekolah_list:
                if sekolah['nama'].strip() == nama_sekolah:
                    id_sekolah = sekolah["id"]
                    nama_sekolah = sekolah["nama"]
                    original_nama = sekolah["nama"]
                    sanitized_nama = original_nama.replace(',', '').strip()  # Hapus koma
                    identifier = sekolah["identifier"]
                    subdomain = base64.b64decode(identifier).decode('utf-8')  # Decode base64 untuk mendapatkan subdomain
                    return id_sekolah, sanitized_nama, subdomain
    print("Tidak ada sekolah yang cocok ditemukan.")
    return None, None, None

# Fungsi untuk mengambil data kelas
def get_kelas(id_sekolah):
    """Mengambil daftar kelas dari sekolah."""
    kelas_url = f"{API_BASE_URL}/{id_sekolah}/datakelas"
    kelas_data = get_json_response(kelas_url)
    return kelas_data if kelas_data is not None else []  # Kembalikan list kosong jika None

# Fungsi untuk mengambil data siswa dalam kelas
def get_siswa(id_sekolah, kelas_id):
    """Mengambil daftar siswa dalam suatu kelas."""
    siswa_url = f"{API_BASE_URL}/v2/{id_sekolah}/listsiswakelas"
    data_siswa = get_json_response(siswa_url, "POST", {"idkelas": kelas_id, "idsiswa": ""})
    return data_siswa

# Fungsi untuk mengambil profil siswa
def get_profil_siswa(id_sekolah, idsiswa):
    """Mengambil profil siswa berdasarkan ID siswa."""
    profil_url = f"{API_BASE_URL}/v2/{id_sekolah}/profilsiswa"
    return get_json_response(profil_url, "POST", {"idsiswa": idsiswa})

# Fungsi untuk login ke subdomain
def login(subdomain):
    """Login ke subdomain untuk mendapatkan sesi login dan menyimpan log cookie."""
    login_url = f'https://{subdomain}.sekolahan.id/login/proses'
    main_url = f'https://{subdomain}.sekolahan.id/'
    login_data = {
        'username': 'superadmin',
        'password': 'sigarantang',
        'submit': ''
    }

    session = requests.Session()
    headers = {'User-Agent': 'Mozilla/5.0'}
    session.get(main_url, headers=headers)

    login_response = session.post(login_url, data=login_data, headers=headers)
    
    if login_response.status_code == 200 and "dashboard" in login_response.text:
        cookies = session.cookies.get_dict()
        cookies_str = "; ".join([f"{k}={v}" for k, v in cookies.items()])
        
        log_to_md(
            subdomain, "login", 
            message=f"Berhasil login. Cookie: {cookies_str}"
        )
        
        return session
    
    log_to_md(subdomain, "login", message="Login gagal.")
    print("Login gagal.")
    return None

# Fungsi untuk scrape data siswa
def scrape_siswa(session, subdomain, idsiswa):
    """Scraping data siswa dari halaman edit."""
    data_siswa_url = f'https://{subdomain}.sekolahan.id/datasiswa/edit/{idsiswa}'
    response = session.get(data_siswa_url)
    if response.status_code == 200:
        soup = BeautifulSoup(response.text, 'html.parser')
        fields = ['nama', 'nik', 'no_kk', 'ayah_nik', 'ibu_nik', 'wali_nik', 'no_kip', 'nm_kip', 'no_kps', 'no_kks', 'tglditerima', 'asalsekolah']
        data = {field: get_input_value(soup, field) for field in fields}
        return data
    else:
        print(f"Gagal mengakses halaman siswa, status code: {response.status_code}")
    return {}

# Tambahkan fungsi ini setelah fungsi scrape_siswa
def download_pdf(session, subdomain, idsiswa, output_path):
    """Download HTML via session, lalu konversi ke PDF"""
    pdf_url = f'https://{subdomain}.sekolahan.id/studentrecord/printdata/{idsiswa}'
    response = session.get(pdf_url)
    
    if response.status_code == 200:
        try:
            # Perbaiki URL relatif dalam HTML
            html_content = response.text.replace('src="/', f'src="https://{subdomain}.sekolahan.id/')
            html_content = html_content.replace('href="/', f'href="https://{subdomain}.sekolahan.id/')
            
            # Simpan cookie session untuk wkhtmltopdf
            cookies = session.cookies.get_dict()
            cookies_str = "; ".join([f"{k}={v}" for k, v in cookies.items()])
            
            options = {
                'quiet': '',
                'print-media-type': '',
                'encoding': 'UTF-8',
                'cookie': [('cookie_name', cookies_str)],  # Kirim cookie
                'custom-header': [
                    ('User-Agent', 'Mozilla/5.0'),  # Samakan dengan header browser
                ],
                'enable-local-file-access': ''  # Izinkan akses file lokal
            }
            
            # Konversi HTML string ke PDF
            pdfkit.from_string(
                html_content, 
                output_path, 
                options=options, 
                configuration=config
            )
            return True
        except Exception as e:
            print(f"Gagal konversi PDF: {str(e)}")
            return False
    else:
        print(f"Gagal akses halaman. Status code: {response.status_code}")
    return False

# Fungsi untuk mendapatkan nilai input dari form
def get_input_value(soup, field):
    """Mengambil nilai dari input berdasarkan nama field."""
    input_tag = soup.find('input', {'name': field})
    return input_tag['value'] if input_tag else None

# TAMBAHKAN FUNGSI INI SETELAH FUNGSI get_input_value
def scrape_profil_sekolah(session, subdomain, sanitized_nama_sekolah):
    """Scraping dan menyimpan profil sekolah ke file txt."""
    try:
        PROFIL_SEKOLAH_URL = f'https://{subdomain}.sekolahan.id/profilsekolah'  # Ganti URL sesuai kebutuhan
        response = session.get(PROFIL_SEKOLAH_URL)
        
        if response.status_code != 200:
            print(f"Gagal mengakses profil sekolah. Status code: {response.status_code}")
            return

        soup = BeautifulSoup(response.text, 'html.parser')
        # Profil Sekolah
        # nama_sekolah = get_input_value(soup, 'pnamasekolah')
        original_nama_sekolah = get_input_value(soup, 'pnamasekolah')
        nss = get_input_value(soup, 'nsssekolah')
        npsn = get_input_value(soup, 'npsnsekolah')
        alamat = get_input_value(soup, 'alamat_sekolah')
        kode_pos = get_input_value(soup, 'kode_possekolah')
        desa_kelurahan = get_input_value(soup, 'desa_kelurahansekolah')
        kecamatan = get_input_value(soup, 'kecamatan_sekolah')
        kabupaten_kota = get_input_value(soup, 'kabupatenkota_sekolah')
        provinsi = get_input_value(soup, 'provinsi_sekolah')
        latitude = get_input_value(soup, 'latitude')
        longitude = get_input_value(soup, 'longitude')

        # Informasi Sekolah
        nomor_telepon = get_input_value(soup, 'nomor_teleponsekolah')
        nomor_fax = get_input_value(soup, 'nomor_faxsekolah')
        email = get_input_value(soup, 'emailsekolah')
        website = get_input_value(soup, 'websitesekolah')
        # HANDLE STATUS KEPEMILIKAN DENGAN CHECK BERLAPIS
        # status_kepemilikan = soup.find('select', {'name': 'status_pemilik'}).find('option', selected=True).text if soup.find('select', {'name': 'status_pemilik'}) else 'Tidak ditemukan'
        status_kepemilikan = 'Tidak ditemukan'
        select_status = soup.find('select', {'name': 'status_pemilik'})
        if select_status:
            selected_option = select_status.find('option', selected=True)
            status_kepemilikan = selected_option.text.strip() if selected_option else ''

        # Kelengkapan Sekolah
        sk_pendirian = get_input_value(soup, 'sk_pendirian_sekolah')
        tgl_sk_pendirian = get_input_value(soup, 'tgl_sk_pendirian_sekolah')
        sk_izin_operasional = get_input_value(soup, 'sk_izin_operasional')
        tgl_sk_izin_operasional = get_input_value(soup, 'tgl_sk_izin_operasional')
        no_rekening = get_input_value(soup, 'no_rekeningsekolah')
        nama_bank = get_input_value(soup, 'nama_banksekolah')
        rekening_atas_nama = get_input_value(soup, 'rekening_atas_nama')

        # Profile Yayasan
        nama_yayasan = get_input_value(soup, 'namayayasan')
        pimpinan_yayasan = get_input_value(soup, 'nama_pimpinan_yayasan')
        alamat_yayasan = get_input_value(soup, 'alamat_yayasan')
        kode_pos_yayasan = get_input_value(soup, 'kode_posyayasan')
        desa_kelurahan_yayasan = get_input_value(soup, 'desa_kelurahanyayasan')
        sk_pendirian_yayasan = get_input_value(soup, 'sk_pendirian_yayasan')
        tgl_sk_pendirian_yayasan = get_input_value(soup, 'tgl_sk_pendirian_yayasan')

            # ========== SIMPAN KE FILE ==========
        output_folder = os.path.join("Data Sekolah", sanitized_nama_sekolah)
        os.makedirs(output_folder, exist_ok=True)
        
        output_path = os.path.join(output_folder, "informasi sekolah.txt")
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write("=== PROFIL SEKOLAH ===\n")
            f.write(f"Nama Sekolah     : {original_nama_sekolah}\n")
            f.write(f"NSS              : {nss}\n")
            f.write(f"NPSN             : {npsn}\n")
            f.write(f"Alamat           : {alamat}\n")
            f.write(f"Kode Pos         : {kode_pos}\n")
            f.write(f"Desa/Kelurahan   : {desa_kelurahan}\n")
            f.write(f"Kecamatan        : {kecamatan}\n")
            f.write(f"Kabupaten/Kota   : {kabupaten_kota}\n")
            f.write(f"Provinsi         : {provinsi}\n")
            f.write(f"Latitude         : {latitude}\n")
            f.write(f"Longitude        : {longitude}\n\n")
            
            f.write("=== INFORMASI KONTAK ===\n")
            f.write(f"Telepon          : {nomor_telepon}\n")
            f.write(f"Fax              : {nomor_fax}\n")
            f.write(f"Email            : {email}\n")
            f.write(f"Website          : {website}\n")
            f.write(f"Status Kepemilikan: {status_kepemilikan}\n\n")
            
            f.write("=== LEGALITAS ===\n")
            f.write(f"SK Pendirian     : {sk_pendirian}\n")
            f.write(f"Tanggal SK       : {tgl_sk_pendirian}\n")
            f.write(f"SK Izin Operasional: {sk_izin_operasional}\n")
            f.write(f"Tanggal SK Izin  : {tgl_sk_izin_operasional}\n\n")
            
            f.write("=== REKENING SEKOLAH ===\n")
            f.write(f"Nomor Rekening   : {no_rekening}\n")
            f.write(f"Nama Bank        : {nama_bank}\n")
            f.write(f"Atas Nama        : {rekening_atas_nama}\n\n")
            
            f.write("=== YAYASAN ===\n")
            f.write(f"Nama Yayasan     : {nama_yayasan}\n")
            f.write(f"Pimpinan         : {pimpinan_yayasan}\n")
            f.write(f"Alamat Yayasan   : {alamat_yayasan}\n")
            f.write(f"Kode Pos Yayasan : {kode_pos_yayasan}\n")
            f.write(f"Desa/Kelurahan   : {desa_kelurahan_yayasan}\n")
            f.write(f"Akte Pendirian   : {sk_pendirian_yayasan}\n")
            f.write(f"Tanggal Akte     : {tgl_sk_pendirian_yayasan}\n")

        print(f"\nProfil sekolah berhasil disimpan di: {output_path}")
    except requests.exceptions.RequestException as e:
        error_msg = f"Gagal akses server: {str(e)}"
        log_to_md(sanitized_nama_sekolah, "sekolah", message=error_msg)
        print(error_msg)
    
    except AttributeError as e:
        error_msg = f"Gagal parsing elemen HTML: {str(e)}"
        log_to_md(sanitized_nama_sekolah, "sekolah", message=error_msg)
        print(error_msg)
    
    except Exception as e:
        error_msg = f"Error tidak terduga: {str(e)}"
        log_to_md(sanitized_nama_sekolah, "sekolah", message=error_msg)
        print(error_msg)

# Download data alumni
def download_alumni(session, subdomain, nama_sekolah):
    alumni_url = f'https://{subdomain}.sekolahan.id/dataalumni/cetakalumni/'
    
    try:
        response = session.get(alumni_url)
        
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            table = soup.find('table')
            
            if not table:
                print("Tabel alumni tidak ditemukan di halaman.")
                return False
            
            # PERBAIKAN 1: Gunakan StringIO untuk handle HTML
            html_content = str(table)
            df = pd.read_html(StringIO(html_content))[0]  # ← Pakai StringIO
            
            # PERBAIKAN 2: Simpan sebagai .xlsx (format modern)
            output_folder = os.path.join("Data Sekolah", nama_sekolah, "Data Siswa")
            os.makedirs(output_folder, exist_ok=True)
            output_path = os.path.join(output_folder, "alumni.xlsx")  # Ganti ke .xlsx

            # Simpan ke Excel dengan styling
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Alumni')
                
                # Ambil worksheet
                workbook = writer.book
                worksheet = writer.sheets['Alumni']
                
                # Tentukan border
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Terapkan border ke semua cell
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Handle kolom NIK dan HP (tambahkan ` di depan)
                for col in df.columns:
                    if 'NIK' in col or 'HP' in col or 'NUPTK' in col or 'NIP' in col:
                        col_idx = df.columns.get_loc(col) + 1  # Kolom Excel dimulai dari 1
                        col_letter = get_column_letter(col_idx)
                        
                        for cell in worksheet[col_letter]:
                            if cell.value:  # Jika ada isinya
                                cell.value = f"'{cell.value}"  # Tambahkan ` di depan
                
                # Auto adjust lebar kolom
                for col in worksheet.columns:
                    max_length = max(len(str(cell.value)) for cell in col)
                    adjusted_width = (max_length + 2) * 1.2  # Tambahkan sedikit padding
                    worksheet.column_dimensions[col[0].column_letter].width = adjusted_width
            
            print(f"\nData alumni berhasil disimpan: {output_path}")
            return True
            
        else:
            print(f"Gagal download alumni. Status code: {response.status_code}")
            return False
            
    except Exception as e:
        print(f"Error saat download alumni: {str(e)}")
        return False

#download data guru
def download_guru(session, subdomain, nama_sekolah):
    """Mendownload data guru dalam format XLSX"""
    guru_url = f'https://{subdomain}.sekolahan.id/dataguru/cetakguru/'
    
    try:
        response = session.get(guru_url)
        
        if response.status_code != 200:
            error_msg = f"Gagal download guru. Status code: {response.status_code}"
            log_to_md(nama_sekolah, "data guru", message=error_msg)
            print(error_msg)
            return False

        # Parsing HTML menggunakan BeautifulSoup
        soup = BeautifulSoup(response.text, 'html.parser')
        table = soup.find('table')  # Mencari tabel data guru
        
        # Cek apakah tabel ditemukan
        if not table:
            error_msg = "Tabel guru tidak ditemukan di halaman."
            log_to_md(nama_sekolah, "data guru", message=error_msg)
            print(error_msg)
            return False
        
        # Konversi tabel HTML ke DataFrame menggunakan pandas
        html_content = str(table)
        df = pd.read_html(StringIO(html_content))[0]
        
        # Membuat folder penyimpanan jika belum ada
        output_folder = os.path.join("Data Sekolah", nama_sekolah, "Data Guru")
        os.makedirs(output_folder, exist_ok=True)
        
        # Path file output
        output_path = os.path.join(output_folder, "Data Guru.xlsx")
        
        # Simpan DataFrame ke file Excel dengan styling
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Guru')
            
            # Ambil worksheet untuk styling
            workbook = writer.book
            worksheet = writer.sheets['Guru']
            
            # Tentukan border untuk sel
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Terapkan border dan alignment ke semua sel
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Handle kolom NIK/HP (tambahkan ` di depan)
            for col in df.columns:
                if any(k in col for k in ['NIK', 'HP', 'Telp', 'No. Telpon']):
                    col_idx = df.columns.get_loc(col) + 1
                    col_letter = get_column_letter(col_idx)
                    
                    for cell in worksheet[col_letter]:
                        if cell.value:
                            cell.value = f"`{cell.value}"
            
            # Auto adjust lebar kolom
            for col in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in col)
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[col[0].column_letter].width = adjusted_width
        
        # Log keberhasilan
        success_msg = "Data Guru.xlsx berhasil di-download"
        log_to_md(nama_sekolah, "data guru", message=success_msg, nama_file="Data Guru.xlsx")
        print(f"\nData guru berhasil disimpan: {output_path}")
        return True
            
    except requests.exceptions.RequestException as e:
        # Log error jika gagal koneksi
        error_msg = f"Gagal koneksi saat download data guru: {str(e)}"
        log_to_md(nama_sekolah, "data guru", message=error_msg)
        print(error_msg)
        return False
    
    except pd.errors.EmptyDataError as e:
        # Log error jika tabel kosong
        error_msg = f"Tabel guru kosong atau tidak valid: {str(e)}"
        log_to_md(nama_sekolah, "data guru", message=error_msg)
        print(error_msg)
        return False
    
    except Exception as e:
        # Log error untuk exception tak terduga
        error_msg = f"Error tidak terduga saat download data guru: {str(e)}"
        log_to_md(nama_sekolah, "data guru", message=error_msg)
        print(error_msg)
        return False

def download_tendik(session, subdomain, nama_sekolah):
    """Mendownload data tendik dalam format XLSX"""
    tendik_url = f'https://{subdomain}.sekolahan.id/datatendik/cetaktendik/'
    
    try:
        response = session.get(tendik_url)
        
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            table = soup.find('table')
            
            if not table:
                error_msg = "Tabel tendik tidak ditemukan di halaman."
                log_to_md(nama_sekolah, "data tendik", message=error_msg)
                return False
            
            html_content = str(table)
            df = pd.read_html(StringIO(html_content))[0]
            
            output_folder = os.path.join("Data Sekolah", nama_sekolah, "data tendik")
            os.makedirs(output_folder, exist_ok=True)
            output_path = os.path.join(output_folder, "data tendik.xlsx")
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Tendik')
                
                workbook = writer.book
                worksheet = writer.sheets['Tendik']
                
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                for col in worksheet.columns:
                    max_length = max(len(str(cell.value)) for cell in col)
                    adjusted_width = (max_length + 2) * 1.2
                    worksheet.column_dimensions[col[0].column_letter].width = adjusted_width
            
            log_to_md(
                nama_sekolah, 
                "data tendik", 
                message="data tendik.xlsx berhasil di-download",
                nama_file="data tendik.xlsx"
            )
            return True
            
        else:
            error_msg = f"Gagal download data tendik. Status code: {response.status_code}"
            log_to_md(nama_sekolah, "data tendik", message=error_msg)
            return False
            
    except Exception as e:
        error_msg = f"Error saat download data tendik: {str(e)}"
        log_to_md(nama_sekolah, "data tendik", message=error_msg)
        return False

def scrape_tendik_profiles(session, subdomain, nama_sekolah):
    """Scraping profil tendik dan mengunduh PDF dari halaman print profil tendik."""
    DATA_TENDIK_URL = f'https://{subdomain}.sekolahan.id/datatendik'

    def scrape_profile_urls(url):
        """Mengambil daftar nama tendik dan URL cetak profil dari halaman tertentu."""
        response = session.get(url)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            table = soup.find('table', class_='table table-striped table-hover')
            tendik_data = []
            
            if table:
                for row in table.find('tbody').find_all('tr'):
                    cols = row.find_all('td')
                    if len(cols) > 1:
                        nama_tendik_tag = cols[2].find('strong')
                        if nama_tendik_tag:
                            nama_tendik = nama_tendik_tag.text.strip()
                            cetak_link = cols[-1].find('a', href=True, string='Cetak Profil')
                            if cetak_link:
                                tendik_data.append((nama_tendik, cetak_link['href']))
                return tendik_data
        error_msg = f"Gagal mengakses halaman: {url}, status code: {response.status_code}"
        log_to_md(nama_sekolah, "data tendik", message=error_msg)
        return []

    try:
        first_page_response = session.get(DATA_TENDIK_URL)
        if first_page_response.status_code == 200:
            soup = BeautifulSoup(first_page_response.text, 'html.parser')
            pagination = soup.find('ul', class_='pagination')
            total_pages = max([int(a.text) for a in pagination.find_all('a') if a.text.isdigit()], default=1)

            log_to_md(nama_sekolah, "data tendik", message=f"Memulai scraping {total_pages} halaman data tendik")
            
            all_tendik_data = []
            for page in range(1, total_pages + 1):
                page_url = f"{DATA_TENDIK_URL}/?halaman={page}"
                log_to_md(nama_sekolah, "data tendik", message=f"Mengakses halaman {page}")
                page_data = scrape_profile_urls(page_url)
                all_tendik_data.extend(page_data)

            if not all_tendik_data:
                log_to_md(nama_sekolah, "data tendik", message="Tidak ada data tendik yang ditemukan")
                return False

            success_count = 0
            for nama_tendik, profil_url in all_tendik_data:
                try:
                    if not profil_url.startswith("http"):
                        profil_url = f"https://{subdomain}.sekolahan.id{profil_url}"

                    response = session.get(profil_url)
                    if response.status_code == 200:
                        output_folder = os.path.join("Data Sekolah", nama_sekolah, "data tendik")
                        os.makedirs(output_folder, exist_ok=True)
                        output_path = os.path.join(output_folder, f"{nama_tendik}.pdf")

                        pdfkit.from_string(
                            response.text, 
                            output_path, 
                            configuration=config,
                            options={'javascript-delay': '3000'}
                        )
                        log_to_md(
                            nama_sekolah, 
                            "data tendik", 
                            message=f"{nama_tendik}.pdf berhasil di-download",
                            nama_file=f"{nama_tendik}.pdf"
                        )
                        success_count += 1
                    else:
                        error_msg = f"Gagal akses profil {nama_tendik} - Status: {response.status_code}"
                        log_to_md(nama_sekolah, "data tendik", message=error_msg)
                except Exception as e:
                    error_msg = f"Gagal menyimpan {nama_tendik}.pdf: {str(e)}"
                    log_to_md(nama_sekolah, "data tendik", message=error_msg)

            log_to_md(
                nama_sekolah, 
                "data tendik", 
                message=f"Total {success_count}/{len(all_tendik_data)} profil tendik berhasil di-download"
            )
            return True
            
        else:
            error_msg = f"Gagal akses halaman pertama data tendik, status code: {first_page_response.status_code}"
            log_to_md(nama_sekolah, "data tendik", message=error_msg)
            return False
            
    except Exception as e:
        error_msg = f"Error sistem saat scraping tendik: {str(e)}"
        log_to_md(nama_sekolah, "data tendik", message=error_msg)
        return False

def scrape_guru(session, subdomain, nama_sekolah):
    """Scraping data guru dan download profil PDF dengan logging detail"""
    DATA_GURU_URL = f'https://{subdomain}.sekolahan.id/dataguru'
    
    def scrape_page(url):
        """Scrape data guru dari halaman tertentu"""
        try:
            response = session.get(url)
            if response.status_code != 200:
                log_to_md(nama_sekolah, "data guru", message=f"Gagal akses halaman: {url}")
                return []
                
            soup = BeautifulSoup(response.text, 'html.parser')
            table = soup.find('table', class_='table table-striped table-hover')
            
            if not table:
                log_to_md(nama_sekolah, "data guru", message=f"Tabel tidak ditemukan di {url}")
                return []
            
            return [
                (
                    cols[2].text.strip(),  # ID Guru
                    cols[3].find('strong').text.strip() if cols[3].find('strong') else cols[3].text.strip()  # Nama Guru
                ) 
                for row in table.find('tbody').find_all('tr') 
                if (cols := row.find_all('td')) and len(cols) > 3
            ]
            
        except Exception as e:
            log_to_md(nama_sekolah, "data guru", message=f"Error scraping halaman: {str(e)}")
            return []

    def download_profil_guru(id_guru, nama_guru):
        """Download profil guru individual dengan logging"""
        try:
            profil_url = f'https://{subdomain}.sekolahan.id/dataguru/cetakprofil/{id_guru}'
            response = session.get(profil_url)
            
            if response.status_code != 200:
                log_to_md(nama_sekolah, "data guru", message=f"Gagal akses profil {nama_guru} - Status: {response.status_code}")
                return False
                
            output_folder = os.path.join("Data Sekolah", nama_sekolah, "data guru")
            os.makedirs(output_folder, exist_ok=True)
            output_path = os.path.join(output_folder, f"{nama_guru}.pdf")

            # Konversi HTML ke PDF
            pdfkit.from_string(
                response.text, 
                output_path, 
                configuration=config,
                options={
                    'javascript-delay': '2000',
                    'enable-local-file-access': ''
                }
            )
            
            # Log sukses
            log_to_md(
                nama_sekolah, 
                "data guru", 
                message=f"{nama_guru}.pdf berhasil di-scrape dan didownload",
                nama_file=f"{nama_guru}.pdf"
            )
            return True
            
        except Exception as e:
            log_to_md(
                nama_sekolah, 
                "data guru", 
                message=f"Gagal download profil {nama_guru} - Error: {str(e)}"
            )
            return False

    try:
        # Ambil jumlah halaman
        first_page = session.get(DATA_GURU_URL)
        soup = BeautifulSoup(first_page.text, 'html.parser')
        pagination = soup.find('ul', class_='pagination')
        total_pages = max([int(a.text) for a in pagination.find_all('a') if a.text.isdigit()], default=1) if pagination else 1

        all_guru = []
        for page in range(1, total_pages + 1):
            page_url = f"{DATA_GURU_URL}/?halaman={page}"
            log_to_md(nama_sekolah, "data guru", message=f"Memulai scraping halaman {page}")
            
            page_data = scrape_page(page_url)
            if not page_data:
                log_to_md(nama_sekolah, "data guru", message=f"Tidak ada data di halaman {page}")
                continue
                
            all_guru.extend(page_data)

        if not all_guru:
            log_to_md(nama_sekolah, "data guru", message="Tidak ada data guru yang ditemukan")
            return False

        # Download semua profil
        success_count = 0
        for id_guru, nama_guru in all_guru:
            log_to_md(nama_sekolah, "data guru", message=f"Memproses {nama_guru}")
            if download_profil_guru(id_guru, nama_guru):
                success_count += 1

        log_to_md(
            nama_sekolah, 
            "data guru", 
            message=f"Total {success_count}/{len(all_guru)} profil guru berhasil di-download"
        )
        return True

    except Exception as e:
        log_to_md(
            nama_sekolah, 
            "data guru", 
            message=f"Error sistem saat scraping guru: {str(e)}"
        )
        return False

def clean_filename(filename):
    """
    Membersihkan nama file dari karakter yang tidak valid dan spasi berlebih
    """
    # Hapus spasi di awal dan akhir
    cleaned = filename.strip()
    # Ganti multiple spaces dengan single space
    cleaned = ' '.join(cleaned.split())
    # Hapus karakter yang tidak valid untuk nama file
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        cleaned = cleaned.replace(char, '')
    return cleaned

# Fungsi untuk menyimpan data siswa ke dalam file terpisah
# Modifikasi fungsi simpan_data_siswa (tambahkan bagian untuk download PDF)
def simpan_data_siswa(nama_sekolah, kelas, siswa_data, id_sekolah, session, subdomain):
    """Menyimpan data siswa dan profilnya ke dalam file terpisah berdasarkan nama siswa dan kelas."""
    try:
        # Bersihkan nama sekolah dan kelas
        clean_nama_sekolah = clean_filename(nama_sekolah)
        clean_kelas = clean_filename(kelas)
        
        # Buat folder utama untuk data siswa
        output_folder = os.path.join("Data Sekolah", clean_nama_sekolah, "Data Siswa", clean_kelas)
        os.makedirs(output_folder, exist_ok=True)
        
        total_siswa = len(siswa_data)
        success_count = 0
        
        # Log awal proses
        log_to_md(
            nama_sekolah, 
            "data siswa", 
            subkategori=clean_kelas, 
            message=f"Memulai proses untuk {total_siswa} siswa"
        )
        
        for siswa in siswa_data:
            try:
                # Ambil profil siswa
                profil = get_profil_siswa(id_sekolah, siswa['idsiswa'])
                scraped_data = scrape_siswa(session, subdomain, siswa['idsiswa'])
                combined_data = {**profil, **scraped_data}
                nama_siswa = clean_filename(profil.get('nama', 'Nama Tidak Ditemukan'))
                
                # Buat folder untuk siswa
                siswa_folder = os.path.join(output_folder, nama_siswa)
                os.makedirs(siswa_folder, exist_ok=True)
                
                # Simpan data siswa ke file .txt
                file_name_txt = f"{nama_siswa}.txt"
                output_file_path_txt = os.path.join(siswa_folder, file_name_txt)
                
                # print(f"Menyimpan data untuk {nama_siswa}")  # Debug print
                
                with open(output_file_path_txt, "w", encoding="utf-8") as file:
                    file.write(f"=== Data Siswa: {nama_siswa} ===\n")
                    file.write(f"ID: {siswa['idsiswa']}\n")
                    file.write(f"Kelas: {kelas}\n\n")
                    
                    # Generate password dari tanggal lahir
                    tgllahir = profil.get("tgllahir", "")
                    password = (
                        datetime.strptime(tgllahir, "%Y-%m-%d").strftime("%y%m%d") 
                        if tgllahir and tgllahir != '0000-00-00' 
                        else "default_password"
                    )
                    file.write(f"Password: {password}\n")
                    
                    # Tulis seluruh data yang ada di combined_data
                    file.write("=== Seluruh Data Siswa ===\n")
                    for key, value in combined_data.items():
                        # Jika value kosong, tuliskan sebagai 'Tidak Ditemukan'
                        if value is None or value == "":
                            value = ""
                        file.write(f"{key.capitalize().replace('_', ' ')}: {value}\n")
                
                # print(f"Data untuk {nama_siswa} berhasil disimpan dalam {file_name_txt}")
                
                # Log sukses simpan .txt
                log_to_md(
                    nama_sekolah, 
                    "data siswa", 
                    subkategori=clean_kelas, 
                    message=f"{nama_siswa}.txt berhasil di-scrape"
                )
                
                # Download PDF profil siswa
                file_name_pdf = f"{nama_siswa}.pdf"
                output_file_path_pdf = os.path.join(siswa_folder, file_name_pdf)
                
                if download_pdf(session, subdomain, siswa['idsiswa'], output_file_path_pdf):
                    # Log sukses download PDF
                    log_to_md(
                        nama_sekolah, 
                        "data siswa", 
                        subkategori=clean_kelas, 
                        message=f"{nama_siswa}.pdf berhasil di-download"
                    )
                    success_count += 1
                else:
                    # Log gagal download PDF
                    log_to_md(
                        nama_sekolah, 
                        "data siswa", 
                        subkategori=clean_kelas, 
                        message=f"Gagal download {nama_siswa}.pdf"
                    )
                
            except Exception as e:
                print(f"Error saat memproses {nama_siswa}: {str(e)}")
                # Log error untuk siswa tertentu
                log_to_md(
                    nama_sekolah, 
                    "data siswa", 
                    subkategori=clean_kelas, 
                    message=f"Error saat memproses {nama_siswa}: {str(e)}"
                )
        
        # Log statistik akhir
        log_to_md(
            nama_sekolah, 
            "data siswa", 
            subkategori=clean_kelas, 
            message=f"Proses selesai. {success_count}/{total_siswa} siswa berhasil diproses"
        )
        
    except Exception as e:
        print(f"Error sistem saat memproses kelas {kelas}: {str(e)}")
        # Log error sistem
        log_to_md(
            nama_sekolah, 
            "data siswa", 
            subkategori=clean_kelas if 'clean_kelas' in locals() else kelas,
            message=f"Error sistem saat memproses kelas {kelas}: {str(e)}"
        )

def log_to_md(nama_sekolah, kategori, subkategori=None, message=None, nama_file=None):
    try:
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        log_path = "log.md"
        
        if os.path.exists(log_path):
            with open(log_path, "r", encoding="utf-8") as f:
                content = f.read()
        else:
            content = ""
        
        new_content = []
        lines = content.split('\n')
        sekolah_found, kategori_found, subkategori_found = False, False, False
        i = 0
        
        while i < len(lines):
            line = lines[i]
            new_content.append(line)
            
            if line.startswith(f"# {nama_sekolah}"):
                sekolah_found = True
            elif sekolah_found and line.startswith(f"## {kategori}"):
                kategori_found = True
            elif kategori_found and subkategori and line.startswith(f"### {subkategori}"):
                subkategori_found = True
            i += 1
        
        if not sekolah_found:
            new_content.extend(["", f"# {nama_sekolah}", ""])
        
        if not kategori_found:
            new_content.extend(["", f"## {kategori}", ""])
        
        if subkategori and not subkategori_found:
            new_content.extend(["", f"### {subkategori}", ""])
        
        log_message = f"{timestamp} - "
        if nama_file:
            log_message += f"{nama_file} "
        log_message += message
        
        new_content.extend([log_message, ""])
        
        with open(log_path, "w", encoding="utf-8") as f:
            f.write("\n".join(new_content))
            
    except Exception as e:
        print(f"Error saat menulis log: {str(e)}")

# Main Program
def main():
    # Baca daftar nama sekolah dari file
    daftar_nama_sekolah = baca_daftar_sekolah("test.txt")
    
    for nama_sekolah in daftar_nama_sekolah:
        print(f"Mencari data untuk sekolah: {nama_sekolah}")
        id_sekolah, nama_sekolah, subdomain = cari_sekolah([nama_sekolah])
        
        if id_sekolah is None:
            print(f"Tidak dapat menemukan data untuk sekolah: {nama_sekolah}")
            continue
        
        kelas_list = get_kelas(id_sekolah) or []  # Pastikan kelas_list adalah list
        
        if not kelas_list:
            print("Tidak ada data kelas yang ditemukan. Melanjutkan ke data lainnya...")
        
        # Login ke subdomain
        session = login(subdomain)
        if session is None:
            continue
        
        # Proses pengambilan data profil sekolah dan alumni
        scrape_profil_sekolah(session, subdomain, nama_sekolah)
        download_alumni(session, subdomain, nama_sekolah)

        # Scrape data guru
        if scrape_guru(session, subdomain, nama_sekolah):
            # Jika ada data guru, jalankan download_guru
            download_guru(session, subdomain, nama_sekolah)
        else:
            print("Tidak ada data guru yang ditemukan. Melewati download_guru.")

        # Scrape data tendik
        if scrape_tendik_profiles(session, subdomain, nama_sekolah):
            # Jika ada data tendik, jalankan download_tendik
            download_tendik(session, subdomain, nama_sekolah)
        else:
            print("Tidak ada data tendik yang ditemukan. Melewati download_tendik.")

        # Loop data siswa hanya jika kelas_list tidak kosong
        for kelas in kelas_list:
            print(f"Mengambil data siswa untuk kelas: {kelas['namakelas']}")
            siswa_list = get_siswa(id_sekolah, kelas["kelasid"])
            
            if siswa_list:
                simpan_data_siswa(nama_sekolah, kelas['namakelas'], siswa_list, id_sekolah, session, subdomain)
                print(f"Data siswa kelas {kelas['namakelas']} berhasil disimpan.")
            else:
                print(f"Tidak ada data siswa di kelas {kelas['namakelas']}")

if __name__ == "__main__":
    main()